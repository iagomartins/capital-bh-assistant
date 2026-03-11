import os
import re
import sys
import unicodedata
import shutil
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Tuple

import customtkinter as ctk
import pandas as pd
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference


# =========================
# Utilidades de dados
# =========================

def _strip_accents(text: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch)
    )


def normalize_col_name(name: str) -> str:
    """
    Normaliza nomes de colunas para facilitar mapeamento de CSV:
    - minúsculas
    - sem acentos
    - remove símbolos e espaços repetidos
    """
    n = _strip_accents(str(name)).lower().strip()
    n = re.sub(r"[\s/_\-]+", " ", n)
    n = re.sub(r"[^a-z0-9 ]+", "", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def money_to_float(value: Any) -> float:
    """
    Sanitização monetária robusta.
    Aceita formatos comuns do BR, inclusive sufixos/observações, ex:
    - "R$ 15.000,00"
    - "R$ 1.040,94 (PG)"
    - "R$431,77"
    - "15000"
    - 15000.0
    Retorna float (0.0 para vazio/não interpretável).
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)) and pd.notna(value):
        return float(value)
    s = str(value).strip()
    if s == "" or s.lower() in {"nan", "none", "null"}:
        return 0.0

    # Mantém apenas dígitos, vírgula, ponto e sinal.
    s = s.replace("\u00a0", " ")  # nbsp
    s = re.sub(r"[^0-9,.\-]+", "", s)

    # Se vier algo como "-1.234,56" (BR) ou "-1234.56" (US), tenta inferir.
    if s.count(",") > 0 and s.count(".") > 0:
        # Assume BR: pontos como milhar, vírgula como decimal.
        s = s.replace(".", "").replace(",", ".")
    elif s.count(",") > 0 and s.count(".") == 0:
        # Só vírgula: assume decimal.
        s = s.replace(",", ".")
    else:
        # Só ponto ou nada: já está ok.
        pass

    # Pode sobrar algo inválido (ex: "-"), protege.
    try:
        return float(s)
    except Exception:
        return 0.0


def safe_read_csv(path: str) -> pd.DataFrame:
    """
    Lê CSV tentando se adaptar a separadores comuns (',' ';') e encoding.
    """
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return pd.read_csv(path, sep=None, engine="python", encoding=enc)
        except Exception:
            continue
    # Última tentativa sem encoding explícito.
    return pd.read_csv(path, sep=None, engine="python")


@dataclass(frozen=True)
class SheetSpec:
    name: str
    columns: List[str]


SHEETS: Dict[str, SheetSpec] = {
    "Vendas": SheetSpec(
        name="Vendas",
        columns=[
            "Mês/Ano",
            "Cliente",
            "Empresa/Administradora",
            "Sinal / Entrada",
            "Crédito da Cota",
            "Valor Total da Venda",
            "Comissão Empresa",
            "Vendedor",
            "Comissão Vendedor",
            "Status",
            "Criado em",
        ],
    ),
    "Gastos": SheetSpec(
        name="Gastos",
        columns=[
            "Mês/Ano",
            "Descrição do Gasto",
            "Setor/Categoria",
            "Valor",
            "Status Pagamento",
            "Criado em",
        ],
    ),
    "Retiradas": SheetSpec(
        name="Retiradas",
        columns=[
            "Mês/Ano",
            "Sócio / Beneficiário",
            "Descrição",
            "Valor",
            "Criado em",
        ],
    ),
}


class FinanceDataStore:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.ensure_workbook()

    def _read_all_sheets(self) -> Dict[str, pd.DataFrame]:
        """
        Lê todas as abas definidas em SHEETS de uma vez,
        garantindo que as colunas estejam corretas.
        """
        self.ensure_workbook()
        dfs: Dict[str, pd.DataFrame] = {}
        for spec in SHEETS.values():
            df = pd.read_excel(self.excel_path, sheet_name=spec.name, engine="openpyxl")
            for col in spec.columns:
                if col not in df.columns:
                    df[col] = pd.NA
            dfs[spec.name] = df[spec.columns]
        return dfs

    def ensure_workbook(self) -> None:
        os.makedirs(os.path.dirname(self.excel_path) or ".", exist_ok=True)
        if not os.path.exists(self.excel_path):
            with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
                for spec in SHEETS.values():
                    pd.DataFrame(columns=spec.columns).to_excel(
                        writer, sheet_name=spec.name, index=False
                    )
            return

        # Garante que abas e colunas existam (sem perder dados).
        try:
            xls = pd.ExcelFile(self.excel_path, engine="openpyxl")
            existing_sheets = set(xls.sheet_names)
        except Exception:
            # Se o arquivo estiver corrompido/inválido, cria novo (melhor do que travar).
            with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
                for spec in SHEETS.values():
                    pd.DataFrame(columns=spec.columns).to_excel(
                        writer, sheet_name=spec.name, index=False
                    )
            return

        changed = False
        data_to_write: Dict[str, pd.DataFrame] = {}
        for spec in SHEETS.values():
            if spec.name not in existing_sheets:
                data_to_write[spec.name] = pd.DataFrame(columns=spec.columns)
                changed = True
                continue
            df = pd.read_excel(self.excel_path, sheet_name=spec.name, engine="openpyxl")
            for col in spec.columns:
                if col not in df.columns:
                    df[col] = pd.NA
                    changed = True
            df = df[spec.columns]
            data_to_write[spec.name] = df

        if changed:
            with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
                for sheet_name, df in data_to_write.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

    def read_sheet(self, sheet_name: str) -> pd.DataFrame:
        dfs = self._read_all_sheets()
        return dfs[sheet_name]

    def append_row(self, sheet_name: str, row: Dict[str, Any]) -> None:
        dfs = self._read_all_sheets()
        spec = SHEETS[sheet_name]
        df = dfs[sheet_name]

        new_row = {col: row.get(col, pd.NA) for col in spec.columns}
        new_row["Criado em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        df2 = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        dfs[sheet_name] = df2

        with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
            for s, df_s in dfs.items():
                df_s.to_excel(writer, sheet_name=s, index=False)

    def append_many(self, sheet_name: str, rows: List[Dict[str, Any]]) -> int:
        if not rows:
            return 0
        dfs = self._read_all_sheets()
        spec = SHEETS[sheet_name]
        df = dfs[sheet_name]
        normalized_rows: List[Dict[str, Any]] = []
        for r in rows:
            new_row = {col: r.get(col, pd.NA) for col in spec.columns}
            new_row["Criado em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            normalized_rows.append(new_row)
        df2 = pd.concat([df, pd.DataFrame(normalized_rows)], ignore_index=True)
        dfs[sheet_name] = df2

        with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
            for s, df_s in dfs.items():
                df_s.to_excel(writer, sheet_name=s, index=False)
        return len(rows)

    def export_with_charts(self, target_path: str) -> None:
        """
        Exporta uma cópia da planilha para outro local e
        adiciona uma aba 'Dashboard_Gráficos' com dois gráficos
        em formato nativo do Excel (pizza de gastos e vendas).
        """
        self.ensure_workbook()
        os.makedirs(os.path.dirname(target_path) or ".", exist_ok=True)
        shutil.copyfile(self.excel_path, target_path)

        wb = load_workbook(target_path)
        if "Dashboard_Gráficos" in wb.sheetnames:
            del wb["Dashboard_Gráficos"]
        ws = wb.create_sheet("Dashboard_Gráficos")

        # Tabelas-resumo usando os mesmos critérios do Dashboard da aplicação.
        vendas = self.read_sheet("Vendas")
        gastos = self.read_sheet("Gastos")

        vendas["Comissão Empresa"] = vendas["Comissão Empresa"].apply(money_to_float)
        gastos["Valor"] = gastos["Valor"].apply(money_to_float)

        vendas_conc = vendas[
            vendas["Status"].astype(str).str.strip().str.lower() == "concluído".lower()
        ]

        by_setor = (
            gastos.groupby(gastos["Setor/Categoria"].astype(str).replace({"": "Outros"}))["Valor"]
            .sum()
            .sort_values(ascending=False)
        )

        by_vendedor = (
            vendas_conc.groupby(
                vendas_conc["Vendedor"].astype(str).replace({"": "Sem vendedor"})
            )["Comissão Empresa"]
            .sum()
            .sort_values(ascending=False)
        )

        row = 1
        ws["A1"] = "Gastos por Setor/Categoria"
        row += 2
        start_gastos = row
        ws.cell(row=row, column=1, value="Setor/Categoria")
        ws.cell(row=row, column=2, value="Valor")
        row += 1
        for idx, (name, val) in enumerate(by_setor.items(), start=row):
            ws.cell(row=idx, column=1, value=str(name))
            ws.cell(row=idx, column=2, value=float(val))
        end_gastos = row + len(by_setor) - 1

        row = end_gastos + 3
        ws.cell(row=row - 1, column=1, value="Ganhos Brutos por Vendedor")
        start_vendas = row
        ws.cell(row=row, column=1, value="Vendedor")
        ws.cell(row=row, column=2, value="Comissão Empresa")
        row += 1
        for idx, (name, val) in enumerate(by_vendedor.items(), start=row):
            ws.cell(row=idx, column=1, value=str(name))
            ws.cell(row=idx, column=2, value=float(val))
        end_vendas = row + len(by_vendedor) - 1

        # Gráfico pizza de Gastos.
        if end_gastos >= start_gastos + 1:
            pie_gastos = PieChart()
            labels = Reference(ws, min_col=1, min_row=start_gastos + 1, max_row=end_gastos)
            data = Reference(ws, min_col=2, min_row=start_gastos, max_row=end_gastos)
            pie_gastos.add_data(data, titles_from_data=True)
            pie_gastos.set_categories(labels)
            pie_gastos.title = "Gastos por Setor/Categoria"
            ws.add_chart(pie_gastos, "E3")

        # Gráfico pizza de Vendas.
        if end_vendas >= start_vendas + 1:
            pie_vendas = PieChart()
            labels = Reference(ws, min_col=1, min_row=start_vendas + 1, max_row=end_vendas)
            data = Reference(ws, min_col=2, min_row=start_vendas, max_row=end_vendas)
            pie_vendas.add_data(data, titles_from_data=True)
            pie_vendas.set_categories(labels)
            pie_vendas.title = "Ganhos Brutos por Vendedor"
            ws.add_chart(pie_vendas, "E20")

        wb.save(target_path)


# =========================
# Mapeamento de CSV -> Campos
# =========================

def build_column_alias_map() -> Dict[str, List[str]]:
    # Sinônimos comuns que aparecem em CSVs exportados/planilhas.
    return {
        # Vendas
        "Mês/Ano": ["mes ano", "mes/ano", "competencia", "periodo", "mês/ano"],
        "Cliente": ["cliente", "nome cliente", "comprador", "nome"],
        "Empresa/Administradora": ["empresa", "administradora", "adm", "consorcio"],
        "Sinal / Entrada": ["sinal", "entrada", "valor entrada", "sinal entrada"],
        "Crédito da Cota": ["credito", "credito cota", "crédito", "valor credito"],
        "Valor Total da Venda": ["total", "valor total", "valor venda", "total venda"],
        "Comissão Empresa": ["comissao empresa", "comissão empresa", "comissao", "comissão"],
        "Vendedor": ["vendedor", "consultor", "responsavel", "corretor"],
        "Comissão Vendedor": ["comissao vendedor", "comissão vendedor", "repasse", "comissao consultor"],
        "Status": ["status", "situacao", "situação"],
        # Gastos
        "Descrição do Gasto": ["descricao", "descrição", "gasto", "historico", "histórico", "descricao gasto"],
        "Setor/Categoria": ["setor", "categoria", "centro custo", "centro de custo"],
        "Valor": ["valor", "valor gasto", "preco", "preço", "total"],
        "Status Pagamento": ["status pagamento", "pagamento", "situacao pagamento", "pago"],
        # Retiradas
        "Sócio / Beneficiário": ["socio", "sócio", "beneficiario", "beneficiário", "pessoa"],
    }


def guess_mapping(df: pd.DataFrame, target_columns: Iterable[str]) -> Dict[str, str]:
    """
    Retorna um dict {campo_alvo: coluna_csv} quando encontra correspondências por alias.
    """
    aliases = build_column_alias_map()
    norm_cols = {normalize_col_name(c): c for c in df.columns}
    mapping: Dict[str, str] = {}

    for tgt in target_columns:
        candidates = [normalize_col_name(tgt)] + [normalize_col_name(a) for a in aliases.get(tgt, [])]
        chosen: Optional[str] = None
        for cand in candidates:
            if cand in norm_cols:
                chosen = norm_cols[cand]
                break
        if chosen is None:
            # Heurística leve: contém palavras-chave
            for ncol, orig in norm_cols.items():
                if all(p in ncol for p in cand.split()):
                    chosen = orig
                    break
        if chosen is not None:
            mapping[tgt] = chosen
    return mapping


def df_rows_to_records(
    df: pd.DataFrame,
    sheet_name: str,
    mapping: Dict[str, str],
) -> List[Dict[str, Any]]:
    spec = SHEETS[sheet_name]
    money_fields = {
        "Vendas": {
            "Sinal / Entrada",
            "Crédito da Cota",
            "Valor Total da Venda",
            "Comissão Empresa",
            "Comissão Vendedor",
        },
        "Gastos": {"Valor"},
        "Retiradas": {"Valor"},
    }[sheet_name]

    records: List[Dict[str, Any]] = []
    for _, row in df.iterrows():
        rec: Dict[str, Any] = {}
        for col in spec.columns:
            if col == "Criado em":
                continue
            src = mapping.get(col)
            if src is None:
                continue
            val = row.get(src, pd.NA)
            if col in money_fields:
                rec[col] = money_to_float(val)
            else:
                rec[col] = "" if pd.isna(val) else str(val).strip()
        records.append(rec)
    return records


# =========================
# UI (CustomTkinter)
# =========================

NIGHT = {
    "bg": "#0b1220",
    "panel": "#0f1a2b",
    "panel_2": "#12213a",
    "text": "#e6eefc",
    "muted": "#a9b7d3",
    "accent": "#2d6cdf",
    "accent_2": "#1f4fb0",
    "danger": "#d85d5d",
    "success": "#2bb673",
}


def fmt_brl(value: float) -> str:
    try:
        s = f"{value:,.2f}"
        s = s.replace(",", "X").replace(".", ",").replace("X", ".")
        return f"R$ {s}"
    except Exception:
        return "R$ 0,00"


class Card(ctk.CTkFrame):
    def __init__(self, master, title: str, value: str = "—", **kwargs):
        super().__init__(master, fg_color=NIGHT["panel"], corner_radius=16, **kwargs)
        self.grid_columnconfigure(0, weight=1)

        self.title_label = ctk.CTkLabel(
            self,
            text=title,
            text_color=NIGHT["muted"],
            font=ctk.CTkFont(size=13, weight="bold"),
            anchor="w",
        )
        self.title_label.grid(row=0, column=0, padx=14, pady=(12, 0), sticky="ew")

        self.value_label = ctk.CTkLabel(
            self,
            text=value,
            text_color=NIGHT["text"],
            font=ctk.CTkFont(size=22, weight="bold"),
            anchor="w",
        )
        self.value_label.grid(row=1, column=0, padx=14, pady=(6, 12), sticky="ew")

    def set_value(self, text: str) -> None:
        self.value_label.configure(text=text)


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("dark-blue")

        self.title("Capital BH Consórcios — Gestão Financeira")
        self.geometry("1180x720")
        self.minsize(1080, 680)
        self.configure(fg_color=NIGHT["bg"])

        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.excel_path = os.path.join(base_dir, "Financeiro_Capital_BH.xlsx")
        self.store = FinanceDataStore(self.excel_path)

        # Layout principal: sidebar + conteúdo
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.sidebar = ctk.CTkFrame(
            self, width=240, fg_color=NIGHT["panel"], corner_radius=0
        )
        self.sidebar.grid(row=0, column=0, sticky="nsw")
        self.sidebar.grid_rowconfigure(10, weight=1)

        self.content = ctk.CTkFrame(self, fg_color=NIGHT["bg"], corner_radius=0)
        self.content.grid(row=0, column=1, sticky="nsew", padx=18, pady=18)
        self.content.grid_rowconfigure(0, weight=1)
        self.content.grid_columnconfigure(0, weight=1)

        self._build_sidebar()

        self.pages: Dict[str, ctk.CTkFrame] = {
            "Dashboard": DashboardPage(self.content, self),
            "Vendas": VendasPage(self.content, self),
            "Gastos": GastosPage(self.content, self),
            "Retiradas": RetiradasPage(self.content, self),
        }

        for p in self.pages.values():
            p.grid(row=0, column=0, sticky="nsew")

        self.show_page("Dashboard")

    def _build_sidebar(self) -> None:
        title = ctk.CTkLabel(
            self.sidebar,
            text="Capital BH\nConsórcios",
            text_color=NIGHT["text"],
            font=ctk.CTkFont(size=20, weight="bold"),
            justify="left",
        )
        title.grid(row=0, column=0, padx=18, pady=(18, 8), sticky="w")

        subtitle = ctk.CTkLabel(
            self.sidebar,
            text="Gestão Financeira",
            text_color=NIGHT["muted"],
            font=ctk.CTkFont(size=13),
        )
        subtitle.grid(row=1, column=0, padx=18, pady=(0, 16), sticky="w")

        def nav_btn(text: str, page: str, row: int) -> None:
            btn = ctk.CTkButton(
                self.sidebar,
                text=text,
                height=42,
                corner_radius=14,
                fg_color=NIGHT["panel_2"],
                hover_color=NIGHT["accent_2"],
                text_color=NIGHT["text"],
                font=ctk.CTkFont(size=14, weight="bold"),
                command=lambda: self.show_page(page),
            )
            btn.grid(row=row, column=0, padx=16, pady=8, sticky="ew")

        nav_btn("Dashboard", "Dashboard", 2)
        nav_btn("Vendas", "Vendas", 3)
        nav_btn("Gastos", "Gastos", 4)
        nav_btn("Retiradas", "Retiradas", 5)

        sep = ctk.CTkFrame(self.sidebar, height=1, fg_color="#20314f")
        sep.grid(row=6, column=0, padx=16, pady=(14, 10), sticky="ew")

        self.path_label = ctk.CTkLabel(
            self.sidebar,
            text=f"Arquivo: {os.path.basename(self.excel_path)}",
            text_color=NIGHT["muted"],
            font=ctk.CTkFont(size=12),
            wraplength=200,
            justify="left",
        )
        self.path_label.grid(row=7, column=0, padx=16, pady=(0, 6), sticky="w")

        open_btn = ctk.CTkButton(
            self.sidebar,
            text="Abrir pasta do arquivo",
            height=32,
            corner_radius=12,
            fg_color=NIGHT["panel_2"],
            hover_color=NIGHT["accent_2"],
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.open_excel_folder,
        )
        open_btn.grid(row=8, column=0, padx=16, pady=(10, 4), sticky="ew")

        export_btn = ctk.CTkButton(
            self.sidebar,
            text="Exportar planilha...",
            height=32,
            corner_radius=12,
            fg_color=NIGHT["accent"],
            hover_color=NIGHT["accent_2"],
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.export_excel,
        )
        export_btn.grid(row=9, column=0, padx=16, pady=(4, 16), sticky="ew")

    def open_excel_folder(self) -> None:
        folder = os.path.dirname(self.excel_path)
        try:
            os.startfile(folder)  # Windows
        except Exception:
            messagebox.showinfo("Info", f"Pasta do arquivo: {folder}")

    def export_excel(self) -> None:
        initial = os.path.join(
            os.path.expanduser("~"),
            "Financeiro_Capital_BH_export.xlsx",
        )
        path = filedialog.asksaveasfilename(
            title="Exportar planilha",
            defaultextension=".xlsx",
            initialfile=os.path.basename(initial),
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        try:
            self.store.export_with_charts(path)
            messagebox.showinfo(
                "Exportação concluída",
                f"Planilha exportada para:\n{path}\n\nInclui aba 'Dashboard_Gráficos' com gráficos em Excel.",
            )
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar planilha.\n\n{e}")

    def show_page(self, name: str) -> None:
        page = self.pages[name]
        page.tkraise()
        # Atualiza dashboard sempre que for exibido.
        if name == "Dashboard":
            cast = self.pages["Dashboard"]
            if isinstance(cast, DashboardPage):
                cast.refresh()

    def notify_data_changed(self) -> None:
        # Chamado após salvar/importar para manter dashboard atualizado.
        dash = self.pages.get("Dashboard")
        if isinstance(dash, DashboardPage):
            dash.refresh()


class PageBase(ctk.CTkFrame):
    def __init__(self, master, app: App):
        super().__init__(master, fg_color=NIGHT["bg"], corner_radius=0)
        self.app = app

    def header(self, title: str, subtitle: str) -> None:
        self.grid_columnconfigure(0, weight=1)
        top = ctk.CTkFrame(self, fg_color=NIGHT["bg"], corner_radius=0)
        top.grid(row=0, column=0, sticky="ew", pady=(0, 14))
        top.grid_columnconfigure(0, weight=1)

        t = ctk.CTkLabel(
            top,
            text=title,
            text_color=NIGHT["text"],
            font=ctk.CTkFont(size=26, weight="bold"),
            anchor="w",
        )
        t.grid(row=0, column=0, sticky="ew")

        s = ctk.CTkLabel(
            top,
            text=subtitle,
            text_color=NIGHT["muted"],
            font=ctk.CTkFont(size=13),
            anchor="w",
        )
        s.grid(row=1, column=0, sticky="ew", pady=(4, 0))


class VendasPage(PageBase):
    def __init__(self, master, app: App):
        super().__init__(master, app)
        self.header("Vendas", "Cadastrar manualmente ou importar histórico via CSV.")

        body = ctk.CTkFrame(self, fg_color=NIGHT["panel"], corner_radius=16)
        body.grid(row=1, column=0, sticky="nsew")
        self.grid_rowconfigure(1, weight=1)
        body.grid_columnconfigure((0, 1), weight=1)

        actions = ctk.CTkFrame(body, fg_color="transparent")
        actions.grid(row=0, column=0, columnspan=2, sticky="ew", padx=14, pady=(14, 6))
        actions.grid_columnconfigure(0, weight=1)

        import_btn = ctk.CTkButton(
            actions,
            text="Importar CSV Histórico",
            height=40,
            corner_radius=14,
            fg_color=NIGHT["accent"],
            hover_color=NIGHT["accent_2"],
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self.import_csv,
        )
        import_btn.grid(row=0, column=1, sticky="e")

        form = ctk.CTkFrame(body, fg_color="transparent", corner_radius=0)
        form.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=14, pady=(0, 14))
        form.grid_columnconfigure((0, 1), weight=1)
        form.grid_rowconfigure((0, 1, 2, 3, 4, 5, 6), weight=1)

        self.entries: Dict[str, Any] = {}

        self._add_entry(form, 0, 0, "Mês/Ano", placeholder="Janeiro/2026")
        self._add_entry(form, 0, 1, "Cliente")
        self._add_entry(form, 1, 0, "Empresa/Administradora", placeholder="Multimarcas, Primo Rossi, Canopus…")
        self._add_entry(form, 1, 1, "Vendedor")
        self._add_entry(form, 2, 0, "Sinal / Entrada", money=True)
        self._add_entry(form, 2, 1, "Crédito da Cota", money=True)
        self._add_entry(form, 3, 0, "Valor Total da Venda", money=True)
        self._add_entry(form, 3, 1, "Comissão Empresa", money=True)
        self._add_entry(form, 4, 0, "Comissão Vendedor", money=True)

        self._add_dropdown(
            form, 4, 1, "Status", ["Concluído", "Cancelado", "Em Análise"], default="Concluído"
        )

        save_btn = ctk.CTkButton(
            form,
            text="Salvar Venda",
            height=44,
            corner_radius=14,
            fg_color=NIGHT["accent"],
            hover_color=NIGHT["accent_2"],
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self.save,
        )
        save_btn.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(16, 8))

    def _add_entry(
        self,
        parent,
        r: int,
        c: int,
        label: str,
        placeholder: str = "",
        money: bool = False,
    ) -> None:
        w = ctk.CTkFrame(parent, fg_color=NIGHT["panel_2"], corner_radius=14)
        w.grid(row=r, column=c, sticky="ew", padx=6, pady=6)
        w.grid_columnconfigure(0, weight=1)
        l = ctk.CTkLabel(
            w,
            text=label,
            text_color=NIGHT["muted"],
            font=ctk.CTkFont(size=12, weight="bold"),
            anchor="w",
        )
        l.grid(row=0, column=0, padx=12, pady=(10, 0), sticky="ew")
        e = ctk.CTkEntry(
            w,
            placeholder_text=placeholder,
            height=36,
            corner_radius=12,
            fg_color="#0b1730",
            text_color=NIGHT["text"],
        )
        e.grid(row=1, column=0, padx=12, pady=(6, 12), sticky="ew")
        self.entries[label] = (e, money)

    def _add_dropdown(
        self,
        parent,
        r: int,
        c: int,
        label: str,
        values: List[str],
        default: str,
    ) -> None:
        w = ctk.CTkFrame(parent, fg_color=NIGHT["panel_2"], corner_radius=14)
        w.grid(row=r, column=c, sticky="ew", padx=6, pady=6)
        w.grid_columnconfigure(0, weight=1)
        l = ctk.CTkLabel(
            w,
            text=label,
            text_color=NIGHT["muted"],
            font=ctk.CTkFont(size=12, weight="bold"),
            anchor="w",
        )
        l.grid(row=0, column=0, padx=12, pady=(10, 0), sticky="ew")
        cb = ctk.CTkOptionMenu(
            w,
            values=values,
            height=36,
            corner_radius=12,
            fg_color="#0b1730",
            button_color=NIGHT["accent_2"],
            button_hover_color=NIGHT["accent"],
            text_color=NIGHT["text"],
        )
        cb.set(default)
        cb.grid(row=1, column=0, padx=12, pady=(6, 12), sticky="ew")
        self.entries[label] = (cb, False)

    def _get_form_data(self) -> Dict[str, Any]:
        out: Dict[str, Any] = {}
        for k, (widget, is_money) in self.entries.items():
            if isinstance(widget, ctk.CTkOptionMenu):
                val = widget.get()
            else:
                val = widget.get().strip()
            if is_money:
                out[k] = money_to_float(val)
            else:
                out[k] = val
        return out

    def _clear_form(self) -> None:
        for _, (widget, _) in self.entries.items():
            if isinstance(widget, ctk.CTkOptionMenu):
                continue
            widget.delete(0, "end")

    def save(self) -> None:
        data = self._get_form_data()
        self.app.store.append_row("Vendas", data)
        self._clear_form()
        self.app.notify_data_changed()
        messagebox.showinfo("Sucesso", "Venda salva no Excel.")

    def import_csv(self) -> None:
        path = filedialog.askopenfilename(
            title="Selecione o CSV de Vendas",
            filetypes=[("CSV", "*.csv"), ("Todos os arquivos", "*.*")],
        )
        if not path:
            return
        try:
            df = safe_read_csv(path)
            mapping = guess_mapping(df, [c for c in SHEETS["Vendas"].columns if c != "Criado em"])
            recs = df_rows_to_records(df, "Vendas", mapping)
            saved = self.app.store.append_many("Vendas", recs)
            self.app.notify_data_changed()
            messagebox.showinfo("Importação concluída", f"Registros importados: {saved}")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao importar CSV.\n\n{e}")


class GastosPage(PageBase):
    DEFAULT_DESC = [
        "CONTADORA",
        "ALUGUEL",
        "INTERNET E TELEFONE",
        "CONTA DE LUZ E CONDOMINIO",
        "ANÚNCIO / MARKETING",
        "LIMPEZA",
        "Hostinger",
        "Microsoft",
        "Escritório",
        "Conta Vivo",
        "Salário Hiago",
        "Salário Sabrina",
        "Passagem Sabrina",
        "Supermercados BH",
    ]
    SETORES = [
        "Custo Fixo/Aluguel",
        "Assinaturas/Software",
        "Salários/Honorários",
        "Tráfego Pago/Marketing",
        "Impostos/Taxas",
        "Estrutura/Limpeza",
        "Outros",
    ]

    def __init__(self, master, app: App):
        super().__init__(master, app)
        self.header("Gastos", "Controle de despesas com categorização e status de pagamento.")

        body = ctk.CTkFrame(self, fg_color=NIGHT["panel"], corner_radius=16)
        body.grid(row=1, column=0, sticky="nsew")
        self.grid_rowconfigure(1, weight=1)

        actions = ctk.CTkFrame(body, fg_color="transparent")
        actions.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 6))
        actions.grid_columnconfigure(0, weight=1)

        import_btn = ctk.CTkButton(
            actions,
            text="Importar CSV de Gastos",
            height=40,
            corner_radius=14,
            fg_color=NIGHT["accent"],
            hover_color=NIGHT["accent_2"],
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self.import_csv,
        )
        import_btn.grid(row=0, column=1, sticky="e")

        form = ctk.CTkFrame(body, fg_color="transparent")
        form.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 14))
        form.grid_columnconfigure((0, 1), weight=1)
        form.grid_rowconfigure((0, 1, 2, 3, 4), weight=1)

        self.month = self._field_entry(form, 0, 0, "Mês/Ano", placeholder="Janeiro/2026")

        self.desc = self._field_combo_editable(
            form, 0, 1, "Descrição do Gasto", values=self.DEFAULT_DESC
        )
        self.setor = self._field_dropdown_required(
            form, 1, 0, "Setor/Categoria", values=self.SETORES, default="Outros"
        )
        self.valor = self._field_entry(form, 1, 1, "Valor", placeholder="R$ 0,00")
        self.status = self._field_dropdown_required(
            form, 2, 0, "Status Pagamento", values=["Pago", "Pendente"], default="Pago"
        )

        save_btn = ctk.CTkButton(
            form,
            text="Salvar Gasto",
            height=44,
            corner_radius=14,
            fg_color=NIGHT["accent"],
            hover_color=NIGHT["accent_2"],
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self.save,
        )
        save_btn.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(16, 8))

    def _field_wrap(self, parent, r: int, c: int, label: str):
        w = ctk.CTkFrame(parent, fg_color=NIGHT["panel_2"], corner_radius=14)
        w.grid(row=r, column=c, sticky="ew", padx=6, pady=6)
        w.grid_columnconfigure(0, weight=1)
        l = ctk.CTkLabel(
            w,
            text=label,
            text_color=NIGHT["muted"],
            font=ctk.CTkFont(size=12, weight="bold"),
            anchor="w",
        )
        l.grid(row=0, column=0, padx=12, pady=(10, 0), sticky="ew")
        return w

    def _field_entry(self, parent, r: int, c: int, label: str, placeholder: str = ""):
        w = self._field_wrap(parent, r, c, label)
        e = ctk.CTkEntry(
            w,
            placeholder_text=placeholder,
            height=36,
            corner_radius=12,
            fg_color="#0b1730",
            text_color=NIGHT["text"],
        )
        e.grid(row=1, column=0, padx=12, pady=(6, 12), sticky="ew")
        return e

    def _field_combo_editable(self, parent, r: int, c: int, label: str, values: List[str]):
        w = self._field_wrap(parent, r, c, label)
        cb = ctk.CTkComboBox(
            w,
            values=values,
            height=36,
            corner_radius=12,
            fg_color="#0b1730",
            border_width=0,
            text_color=NIGHT["text"],
            button_color=NIGHT["accent_2"],
            button_hover_color=NIGHT["accent"],
            state="normal",  # permite digitação livre
        )
        if values:
            cb.set(values[0])
        cb.grid(row=1, column=0, padx=12, pady=(6, 12), sticky="ew")
        return cb

    def _field_dropdown_required(self, parent, r: int, c: int, label: str, values: List[str], default: str):
        w = self._field_wrap(parent, r, c, label)
        om = ctk.CTkOptionMenu(
            w,
            values=values,
            height=36,
            corner_radius=12,
            fg_color="#0b1730",
            button_color=NIGHT["accent_2"],
            button_hover_color=NIGHT["accent"],
            text_color=NIGHT["text"],
        )
        om.set(default)
        om.grid(row=1, column=0, padx=12, pady=(6, 12), sticky="ew")
        return om

    def save(self) -> None:
        setor = self.setor.get().strip()
        if not setor:
            messagebox.showerror("Obrigatório", "Selecione um Setor/Categoria.")
            return
        row = {
            "Mês/Ano": self.month.get().strip(),
            "Descrição do Gasto": self.desc.get().strip(),
            "Setor/Categoria": setor,
            "Valor": money_to_float(self.valor.get().strip()),
            "Status Pagamento": self.status.get().strip(),
        }
        self.app.store.append_row("Gastos", row)
        self.month.delete(0, "end")
        self.valor.delete(0, "end")
        # Mantém descrição/setor/status para acelerar rotina.
        self.app.notify_data_changed()
        messagebox.showinfo("Sucesso", "Gasto salvo no Excel.")

    def import_csv(self) -> None:
        path = filedialog.askopenfilename(
            title="Selecione o CSV de Gastos",
            filetypes=[("CSV", "*.csv"), ("Todos os arquivos", "*.*")],
        )
        if not path:
            return
        try:
            df = safe_read_csv(path)
            mapping = guess_mapping(df, [c for c in SHEETS["Gastos"].columns if c != "Criado em"])
            recs = df_rows_to_records(df, "Gastos", mapping)
            saved = self.app.store.append_many("Gastos", recs)
            self.app.notify_data_changed()
            messagebox.showinfo("Importação concluída", f"Registros importados: {saved}")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao importar CSV.\n\n{e}")


class RetiradasPage(PageBase):
    def __init__(self, master, app: App):
        super().__init__(master, app)
        self.header("Retiradas", "Controle de retiradas de sócios/beneficiários.")

        body = ctk.CTkFrame(self, fg_color=NIGHT["panel"], corner_radius=16)
        body.grid(row=1, column=0, sticky="nsew")
        self.grid_rowconfigure(1, weight=1)

        actions = ctk.CTkFrame(body, fg_color="transparent")
        actions.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 6))
        actions.grid_columnconfigure(0, weight=1)

        import_btn = ctk.CTkButton(
            actions,
            text="Importar CSV de Retiradas",
            height=40,
            corner_radius=14,
            fg_color=NIGHT["accent"],
            hover_color=NIGHT["accent_2"],
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self.import_csv,
        )
        import_btn.grid(row=0, column=1, sticky="e")

        form = ctk.CTkFrame(body, fg_color="transparent")
        form.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 14))
        form.grid_columnconfigure((0, 1), weight=1)
        form.grid_rowconfigure((0, 1, 2, 3), weight=1)

        self.month = self._field_entry(form, 0, 0, "Mês/Ano", placeholder="Janeiro/2026")
        self.socio = self._field_dropdown(
            form, 0, 1, "Sócio / Beneficiário", ["Alana", "Gomes", "Sabrina", "Outro"], default="Alana"
        )
        self.desc = self._field_entry(form, 1, 0, "Descrição")
        self.valor = self._field_entry(form, 1, 1, "Valor", placeholder="R$ 0,00")

        save_btn = ctk.CTkButton(
            form,
            text="Salvar Retirada",
            height=44,
            corner_radius=14,
            fg_color=NIGHT["accent"],
            hover_color=NIGHT["accent_2"],
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self.save,
        )
        save_btn.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(16, 8))

    def _wrap(self, parent, r: int, c: int, label: str):
        w = ctk.CTkFrame(parent, fg_color=NIGHT["panel_2"], corner_radius=14)
        w.grid(row=r, column=c, sticky="ew", padx=6, pady=6)
        w.grid_columnconfigure(0, weight=1)
        l = ctk.CTkLabel(
            w,
            text=label,
            text_color=NIGHT["muted"],
            font=ctk.CTkFont(size=12, weight="bold"),
            anchor="w",
        )
        l.grid(row=0, column=0, padx=12, pady=(10, 0), sticky="ew")
        return w

    def _field_entry(self, parent, r: int, c: int, label: str, placeholder: str = ""):
        w = self._wrap(parent, r, c, label)
        e = ctk.CTkEntry(
            w,
            placeholder_text=placeholder,
            height=36,
            corner_radius=12,
            fg_color="#0b1730",
            text_color=NIGHT["text"],
        )
        e.grid(row=1, column=0, padx=12, pady=(6, 12), sticky="ew")
        return e

    def _field_dropdown(self, parent, r: int, c: int, label: str, values: List[str], default: str):
        w = self._wrap(parent, r, c, label)
        om = ctk.CTkOptionMenu(
            w,
            values=values,
            height=36,
            corner_radius=12,
            fg_color="#0b1730",
            button_color=NIGHT["accent_2"],
            button_hover_color=NIGHT["accent"],
            text_color=NIGHT["text"],
        )
        om.set(default)
        om.grid(row=1, column=0, padx=12, pady=(6, 12), sticky="ew")
        return om

    def save(self) -> None:
        row = {
            "Mês/Ano": self.month.get().strip(),
            "Sócio / Beneficiário": self.socio.get().strip(),
            "Descrição": self.desc.get().strip(),
            "Valor": money_to_float(self.valor.get().strip()),
        }
        self.app.store.append_row("Retiradas", row)
        self.month.delete(0, "end")
        self.desc.delete(0, "end")
        self.valor.delete(0, "end")
        self.app.notify_data_changed()
        messagebox.showinfo("Sucesso", "Retirada salva no Excel.")

    def import_csv(self) -> None:
        path = filedialog.askopenfilename(
            title="Selecione o CSV de Retiradas",
            filetypes=[("CSV", "*.csv"), ("Todos os arquivos", "*.*")],
        )
        if not path:
            return
        try:
            df = safe_read_csv(path)
            mapping = guess_mapping(df, [c for c in SHEETS["Retiradas"].columns if c != "Criado em"])
            recs = df_rows_to_records(df, "Retiradas", mapping)
            saved = self.app.store.append_many("Retiradas", recs)
            self.app.notify_data_changed()
            messagebox.showinfo("Importação concluída", f"Registros importados: {saved}")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao importar CSV.\n\n{e}")


class DashboardPage(PageBase):
    def __init__(self, master, app: App):
        super().__init__(master, app)
        self.header("Dashboard", "Resumo global e gráficos atualizados em tempo real a partir do Excel.")

        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        cards = ctk.CTkFrame(self, fg_color="transparent")
        cards.grid(row=1, column=0, sticky="ew", pady=(0, 14))
        cards.grid_columnconfigure((0, 1, 2, 3), weight=1, uniform="c")

        self.card_ganhos = Card(cards, "Ganhos Brutos")
        self.card_ganhos.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        self.card_gastos = Card(cards, "Total de Gastos")
        self.card_gastos.grid(row=0, column=1, sticky="ew", padx=10)

        self.card_liquido = Card(cards, "Ganhos Líquidos")
        self.card_liquido.grid(row=0, column=2, sticky="ew", padx=10)

        self.card_cac = Card(cards, "CAC Estimado")
        self.card_cac.grid(row=0, column=3, sticky="ew", padx=(10, 0))

        bottom = ctk.CTkFrame(self, fg_color=NIGHT["panel"], corner_radius=16)
        bottom.grid(row=2, column=0, sticky="nsew")
        self.grid_rowconfigure(2, weight=1)
        bottom.grid_rowconfigure(0, weight=1)
        bottom.grid_columnconfigure((0, 1), weight=1, uniform="g")

        self.chart_left = ctk.CTkFrame(bottom, fg_color=NIGHT["panel_2"], corner_radius=16)
        self.chart_left.grid(row=0, column=0, sticky="nsew", padx=(14, 7), pady=14)
        self.chart_left.grid_rowconfigure(1, weight=1)
        self.chart_left.grid_columnconfigure(0, weight=1)

        self.chart_right = ctk.CTkFrame(bottom, fg_color=NIGHT["panel_2"], corner_radius=16)
        self.chart_right.grid(row=0, column=1, sticky="nsew", padx=(7, 14), pady=14)
        self.chart_right.grid_rowconfigure(1, weight=1)
        self.chart_right.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.chart_left,
            text="Gastos por Setor/Categoria",
            text_color=NIGHT["text"],
            font=ctk.CTkFont(size=14, weight="bold"),
        ).grid(row=0, column=0, padx=14, pady=(12, 6), sticky="w")

        ctk.CTkLabel(
            self.chart_right,
            text="Ganhos Brutos por Vendedor",
            text_color=NIGHT["text"],
            font=ctk.CTkFont(size=14, weight="bold"),
        ).grid(row=0, column=0, padx=14, pady=(12, 6), sticky="w")

        self._canvas_left: Optional[FigureCanvasTkAgg] = None
        self._canvas_right: Optional[FigureCanvasTkAgg] = None

    def refresh(self) -> None:
        vendas = self.app.store.read_sheet("Vendas")
        gastos = self.app.store.read_sheet("Gastos")
        retiradas = self.app.store.read_sheet("Retiradas")

        # Converte campos monetários (caso planilha tenha strings antigas).
        for col in [
            "Sinal / Entrada",
            "Crédito da Cota",
            "Valor Total da Venda",
            "Comissão Empresa",
            "Comissão Vendedor",
        ]:
            vendas[col] = vendas[col].apply(money_to_float)
        gastos["Valor"] = gastos["Valor"].apply(money_to_float)
        retiradas["Valor"] = retiradas["Valor"].apply(money_to_float)

        vendas_conc = vendas[vendas["Status"].astype(str).str.strip().str.lower() == "concluído".lower()]
        ganhos_brutos = float(vendas_conc["Comissão Empresa"].sum())
        total_gastos = float(gastos["Valor"].sum())
        total_retiradas = float(retiradas["Valor"].sum())
        ganhos_liquidos = ganhos_brutos - total_gastos - total_retiradas

        qtd_vendas_conc = int(len(vendas_conc))
        trafego = gastos[
            gastos["Setor/Categoria"].astype(str).str.strip().str.lower()
            == "tráfego pago/marketing".lower()
        ]["Valor"].sum()
        cac = float(trafego) / qtd_vendas_conc if qtd_vendas_conc > 0 else 0.0

        self.card_ganhos.set_value(fmt_brl(ganhos_brutos))
        self.card_gastos.set_value(fmt_brl(total_gastos))
        self.card_liquido.set_value(fmt_brl(ganhos_liquidos))
        self.card_cac.set_value(fmt_brl(cac))

        # Gráfico 1: pizza de gastos por setor
        by_setor = (
            gastos.groupby(gastos["Setor/Categoria"].astype(str).replace({"": "Outros"}))["Valor"]
            .sum()
            .sort_values(ascending=False)
        )
        self._render_pie(
            parent=self.chart_left,
            which="left",
            series=by_setor,
            empty_text="Sem dados de gastos.",
        )

        # Gráfico 2: pizza ganhos brutos por vendedor (vendas concluídas)
        by_vendedor = (
            vendas_conc.groupby(vendas_conc["Vendedor"].astype(str).replace({"": "Sem vendedor"}))[
                "Comissão Empresa"
            ]
            .sum()
            .sort_values(ascending=False)
        )
        self._render_pie(
            parent=self.chart_right,
            which="right",
            series=by_vendedor,
            empty_text="Sem vendas concluídas.",
        )

    def _render_pie(self, parent, which: str, series: pd.Series, empty_text: str) -> None:
        # Limpa canvas anterior (evita sobreposição e consumo de memória).
        if which == "left" and self._canvas_left is not None:
            self._canvas_left.get_tk_widget().destroy()
            self._canvas_left = None
        if which == "right" and self._canvas_right is not None:
            self._canvas_right.get_tk_widget().destroy()
            self._canvas_right = None

        fig = Figure(figsize=(5, 3), dpi=110)
        bg = NIGHT["panel_2"]
        fig.patch.set_facecolor(bg)
        ax = fig.add_subplot(111)
        ax.set_facecolor(bg)

        if series is None or len(series) == 0 or float(series.sum()) == 0.0:
            ax.text(
                0.5,
                0.5,
                empty_text,
                ha="center",
                va="center",
                color=NIGHT["muted"],
                fontsize=12,
                fontweight="bold",
            )
            ax.set_axis_off()
        else:
            labels = list(series.index.astype(str))
            values = list(series.values)

            # Paleta azul (harmoniza com dark mode).
            base_colors = ["#2d6cdf", "#3b82f6", "#1f4fb0", "#60a5fa", "#93c5fd", "#2563eb", "#0ea5e9"]
            colors = [base_colors[i % len(base_colors)] for i in range(len(values))]

            ax.pie(
                values,
                labels=labels,
                autopct=lambda p: f"{p:.0f}%" if p >= 6 else "",
                startangle=90,
                colors=colors,
                textprops={"color": NIGHT["text"], "fontsize": 10},
                wedgeprops={"linewidth": 1, "edgecolor": bg},
            )
            ax.axis("equal")

        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas.get_tk_widget().grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 14))

        if which == "left":
            self._canvas_left = canvas
        else:
            self._canvas_right = canvas


def main() -> None:
    try:
        app = App()
        app.mainloop()
    except Exception as e:
        # Evita “sumir” em execução double-click; mostra erro amigável.
        messagebox.showerror("Erro inesperado", f"Ocorreu um erro:\n\n{e}")
        raise


if __name__ == "__main__":
    main()

